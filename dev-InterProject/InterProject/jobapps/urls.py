import csv
import os
import random
from django.utils import timezone
from datetime import timedelta
import logging
import pandas as pd
import requests
from django.contrib import admin
from django.urls import path, include, re_path
from django.views.generic import RedirectView  # 添加这行导入
from openpyxl.styles import NamedStyle
from rest_framework.documentation import include_docs_urls
from rest_framework import permissions
from drf_yasg.views import get_schema_view
from drf_yasg import openapi
from apscheduler.schedulers.background import BackgroundScheduler
from django_apscheduler.jobstores import DjangoJobStore, register_job
import runpy
from datetime import datetime
from openpyxl import Workbook
from collections import defaultdict

from jobapps.job.models import Job
from jobapps.middle.models import ClickJob, CompanyCode
from jobapps.student.models import Student
from jobapps.user.models import User
from django.http import HttpResponse
# schema_view = get_schema_view(
#     openapi.Info(
#         title="Snippets API",
#         default_version='v1',
#         description="Test description",
#         terms_of_service="https://www.google.com/policies/terms/",
#         contact=openapi.Contact(email="contact@snippets.local"),
#         license=openapi.License(name="BSD License"),
#     ),
#     # public 表示文档完全公开, 无需针对用户鉴权
#     public=True,
#     # 可以传递 drf 的 BasePermission
#     permission_classes=[permissions.AllowAny],
# )

urlpatterns = [
    # 根路径重定向到API文档
    path('', RedirectView.as_view(url='/docs/', permanent=True)),
    
    # path('api-auth/', include('rest_framework.urls')),  # DRF登录退出
    # path("admin/", admin.site.urls),  # 后台管理
    # path('docs/', include_docs_urls(title='API文档', description='API文档')),  # api文档
    path('api/user/', include('jobapps.user.urls')),  # 用户相关
    path('api/company/', include('jobapps.company.urls')),  # 企业相关
    path('api/job/', include('jobapps.job.urls')),  # 职位相关
    path('api/work/', include('jobapps.work.urls')),  # 职位详情相关
    path('api/student/', include('jobapps.student.urls')),  # 学生相关
    path('api/resume/', include('jobapps.resume.urls')),  # 学生相关
    path('api/recruit/', include('jobapps.recruit.urls')),  # 学生相关
    # # drf_yasg
    # re_path(r'^swagger(?P<format>\.json|\.yaml)$', schema_view.without_ui(cache_timeout=0), name='schema-spec'),
    # re_path(r'^swagger/$', schema_view.with_ui('swagger', cache_timeout=0), name='schema-swagger-ui'),
    # re_path(r'^redoc/$', schema_view.with_ui('redoc', cache_timeout=0), name='schema-redoc'),
]


# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# # 重新运行模型
# print("===开始重新运行模型===")
# # 增量训练
# runpy.run_path('JobRec/DPGNN/model/Incremental_training_xdu.py')
#
# # 合并数据集
# runpy.run_path('JobRec/dataset/merge_dataset_xdu.py')
#
# # 重新训练模型
# runpy.run_path('JobRec/DPGNN/model/train.py')


"""
把work的数据和job的数据合并
放到job这张表中
"""


def combineWorkAndJob():
    from work.models import Work

    print("职位数据迁移开始")
    # 获取所有 Work 对象
    works = Work.objects.all()

    # 遍历每个 Work 对象
    for work in works:
        # 创建 Job 对象
        job = Job()

        # 根据字段映射关系填充 Job 字段
        job.SJDWMC = work.companyName or ""  # 公司名
        job.GZZWLBMC = work.workName or ""  # 职位名
        job.major = work.major or ""  # 需求专业
        job.salary = work.salary or ""  # 薪资
        job.degree = work.degree or ""  # 学历
        job.workNature = work.workNature or ""  # 工作性质
        job.desc = work.desc or ""  # 职位描述

        # 其他字段用空字符串或默认值填充
        job.DWXZMC = ""  # 公司性质
        job.DWHYMC = ""  # 公司所属行业
        job.DWSZDDM = ""  # 工作地址

        # 假设 Company 模型可以通过某种方式关联，这里用 None 或默认值
        job.company = None  # 如果需要关联公司，需补充逻辑

        # 保存 Job 对象
        job.save()
        # 设置职位编号
        job.DWZZJGDM = job.id
        job.save()
        print(job.id)

    print("迁移完成！")


"""
同步学生数据
同时将学生数据保存到冷启动项目中
"""


def syncStudentData():
    print("开始同步本科生数据")
    key = '20240603380417974952305645404164571'
    secret = '567c49a47e435a09a27f33b6d18c91982886628a'
    access_token = ''
    per_page = 10000
    page = 1

    while True:
        # 获取token
        response = requests.get(f'https://dmp.xidian.edu.cn:8771/open_api/authentication/get_access_token?'
                                f'key={key}&secret={secret}')

        if response.status_code == 200:
            data = response.json()
            access_token = data.get('result').get('access_token')

        # 使用token发送请求
        response = requests.get(f'https://dmp.xidian.edu.cn:8771/open_api/customization/tgxxsbksjbxx/full?'
                                f'access_token={access_token}&per_page={per_page}&page={page}')
        if response.status_code == 200:
            data = response.json()
            max_page = data.get('result').get('max_page')
            print(page, max_page)

            students = data.get('result').get('data')

            for student in students:
                if Student.objects.filter(SID=student.get('XH')).exists():
                    continue

                saveStudent = Student()

                saveStudent.SID = student.get('XH')
                saveStudent.REALNAME = student.get('XM')
                saveStudent.DEPARTMENT = student.get('XSM')
                saveStudent.MAJOR = student.get('ZYM')
                saveStudent.ZYFX = student.get('ZYFXMC')
                saveStudent.XBMC = student.get('XB')
                saveStudent.BIRTHDAY = student.get('CSRQ')
                saveStudent.XLMC = "本科生"
                saveStudent.XXXSMC = "全日制"

                saveStudent.save()

            if page == max_page:
                break
            else:
                page = page + 1

    print("开始同步研究生数据")
    page = 1

    while True:
        # 获取token
        response = requests.get(f'https://dmp.xidian.edu.cn:8771/open_api/authentication/get_access_token?'
                                f'key={key}&secret={secret}')

        if response.status_code == 200:
            data = response.json()
            access_token = data.get('result').get('access_token')

        # 使用token发送请求
        response = requests.get(f'https://dmp.xidian.edu.cn:8771/open_api/customization/tgxxsyjsjbxx/full?'
                                f'access_token={access_token}&per_page={per_page}&page={page}')
        if response.status_code == 200:
            data = response.json()
            max_page = data.get('result').get('max_page')
            print(page, max_page)

            students = data.get('result').get('data')

            for student in students:
                if Student.objects.filter(SID=student.get('XH')).exists():
                    continue

                saveStudent = Student()

                saveStudent.SID = student.get('XH')
                saveStudent.REALNAME = student.get('XM')
                saveStudent.MAJOR = student.get('YJFX')
                saveStudent.ZYFX = student.get('YJFX')
                saveStudent.BIRTHDAY = student.get('CSRQ')
                saveStudent.XLMC = "研究生"
                saveStudent.XXXSMC = "全日制"
                # 性别
                if student.get('XBDM') == "1":
                    saveStudent.XBMC = "男"
                else:
                    saveStudent.XBMC = "女"

                saveStudent.save()

            if page == max_page:
                break
            else:
                page = page + 1

    # 调整学生数据,适应模型
    students = Student.objects.filter(MAJOR="电子信息类 （电子科学、信息物理、空天信息）")
    for student in students:
        student.MAJOR = "计算机技术"
        student.save()


# 将数据库中学生的数据保存为文件的形式
def studentToXlsx():
    base_dir = "data/xdu/"
    file_name = "xdu_dataset_user.xlsx"
    file_path = os.path.join(base_dir, file_name)

    students = Student.objects.all()

    fieldnames = [
        "SID", "REALNAME", "GRADE", "DEPARTMENT", "MAJOR",
        "ZYFX", "XBMC", "ZXWYYZMC", "BIRTHDAY", "XLMC", "XXXSMC",
        "JTDZ", "GZZWLBMC"
    ]

    # 创建一个工作簿和工作表
    workbook = Workbook()
    worksheet = workbook.active

    # 写入表头
    worksheet.append(fieldnames)

    # 写入学生数据
    for student in students:
        # 获取学生信息并处理空值
        row = [
            student.SID or "",
            student.REALNAME or "",
            student.GRADE or "",
            student.DEPARTMENT or "",
            student.MAJOR or "",
            student.ZYFX or "",
            student.XBMC or "",
            student.ZXWYYZMC or "",
            student.BIRTHDAY or "",
            student.XLMC or "",
            student.XXXSMC or "",
            student.JTDZ or "",
            student.GZZWLBMC or "",
        ]
        worksheet.append(row)

    # 保存文件
    workbook.save(file_path)

    print(f"数据已成功导出到文件: {file_path}")


"""
为所有学生创建用户账号
"""


def createUsersForStudents():
    students = Student.objects.all()
    for student in students:  # 遍历所有学生
        username = student.SID  # 用户名默认为学生学号
        password = '123456'  # 密码默认设为123456
        if not User.objects.filter(username=username).exists():  # 若用户不存在
            user = User.objects.create_user(username=username, password=password)
            user.save()
            # 进行外键关联
            student.user = user
            student.save()


"""
同步职位数据
同时将职位数据保存到冷启动项目中
"""


def syncJobData():
    print("开始同步职位数据")
    key = '20240603380417974952305645404164571'
    secret = '567c49a47e435a09a27f33b6d18c91982886628a'
    access_token = ''
    per_page = 10000
    page = 1
    num = 0

    while True:
        # 获取token
        response = requests.get(f'https://dmp.xidian.edu.cn:8771/open_api/authentication/get_access_token?'
                                f'key={key}&secret={secret}')

        if response.status_code == 200:
            data = response.json()
            access_token = data.get('result').get('access_token')

        # 使用token发送请求
        response = requests.get(f'https://dmp.xidian.edu.cn:8771/open_api/customization/tjyxtxxdviewjobinfo/full?'
                                f'access_token={access_token}&per_page={per_page}&page={page}')
        if response.status_code == 200:
            data = response.json()
            max_page = data.get('result').get('max_page')
            print(page, max_page)

            jobs = data.get('result').get('data')

            for job in jobs:
                # 仅保存在有效期的职位
                valid_time = job.get('VALIDTIME')
                # 将 Unix 时间戳转换为 datetime 对象
                valid_datetime = datetime.fromtimestamp(int(valid_time))

                if Job.objects.filter(DWZZJGDM=job.get('JOBID')).exists():
                    updateJob = Job.objects.get(DWZZJGDM=job.get('JOBID'))
                    # 更新企业编号
                    updateJob.COMPANYID = job.get('COMPANYID')

                    if valid_datetime < timezone.now():
                        updateJob.delete()
                    else:
                        updateJob.save()
                    continue

                if valid_datetime < timezone.now():
                    continue
                if valid_datetime >= timezone.now():
                    num += 1

                saveJob = Job()

                saveJob.DWZZJGDM = job.get('JOBID')
                saveJob.SJDWMC = job.get('COMPANY_NAME')
                saveJob.GZZWLBMC = job.get('JOB_NAME')
                saveJob.DWSZDDM = job.get('ADDRESS')
                saveJob.major = job.get('D_MAJOR')
                saveJob.salary = job.get('D_SALARY')
                saveJob.desc = job.get('DESCRIPTION')
                # 保存企业id
                saveJob.COMPANYID = job.get('COMPANYID')
                # 转换时间格式并保存
                date_line = job.get('DATELINE')
                saveJob.create_time = datetime.fromtimestamp(int(date_line))

                saveJob.save()

            if page == max_page:
                break
            else:
                page = page + 1

    print("有效的职位数量:", num)


# 将数据库中的职位数据保存为文件的形式
def jobToXlsx():
    base_dir = "data/xdu/"
    file_name = "xdu_dataset_job.xlsx"
    file_path = os.path.join(base_dir, file_name)

    jobs = Job.objects.all()

    fieldnames = [
        "DWZZJGDM", "SJDWMC", "DWXZMC", "DWHYMC", "GZZWLBMC", "DWSZDDM"]

    # 创建一个工作簿和工作表
    workbook = Workbook()
    worksheet = workbook.active

    # 写入表头
    worksheet.append(fieldnames)

    # 写入学生数据
    for job in jobs:
        # 获取学生信息并处理空值
        row = [
            job.DWZZJGDM or "",
            job.SJDWMC or "",
            # 公司性质
            # job.DWXZMC or "",
            "民营企业",
            # 公司所属行业
            # job.DWHYMC or "",
            "默认",
            job.GZZWLBMC or "",
            job.DWSZDDM or "",
        ]
        worksheet.append(row)

    # 保存文件
    workbook.save(file_path)

    print(f"数据已成功导出到文件: {file_path}")


def sync_job_data_and_analyze_update_time():
    print("开始分析职位数据")
    key = '20240603380417974952305645404164571'
    secret = '567c49a47e435a09a27f33b6d18c91982886628a'
    per_page = 10000
    page = 1

    num = 1

    while True:
        # 获取token
        response = requests.get(
            f'https://dmp.xidian.edu.cn:8771/open_api/authentication/get_access_token?key={key}&secret={secret}')
        if response.status_code == 200:
            data = response.json()
            access_token = data.get('result').get('access_token')
        else:
            print("获取 Access Token 失败")
            break

        # 使用token发送请求
        response = requests.get(
            f'https://dmp.xidian.edu.cn:8771/open_api/customization/tjyxtxxdviewjobinfo/full?access_token={access_token}&per_page={per_page}&page={page}')
        if response.status_code == 200:
            data = response.json()
            max_page = data.get('result').get('max_page')
            print(f"当前页码: {page}, 总页码: {max_page}")

            jobs = data.get('result').get('data')

            for job in jobs:
                # 假设 UPDATETIME 是一个日期时间字符串
                valid_time = job.get('VALIDTIME')
                valid_datetime = datetime.fromtimestamp(int(valid_time))

                if valid_datetime >= timezone.now():
                    num = num + 1
                    continue

            # 如果当前页码等于最大页码，结束循环
            if page == max_page:
                break
            else:
                page += 1

    print("有效的职位数量:", num)  # 2268


# 初始化交互记录
def createAction():
    base_dir = "data/xdu/"
    file_name = "xdu_dataset_action.xlsx"
    file_path = os.path.join(base_dir, file_name)

    jobs = Job.objects.all()[:100]
    students = Student.objects.all()

    fieldnames = [
        "SID", "DWZZJGDM", "satisfied"]

    # 创建一个工作簿和工作表
    workbook = Workbook()
    worksheet = workbook.active

    # 写入表头
    worksheet.append(fieldnames)

    for job in jobs:
        row = [
            str(random.choice(students).SID) if random.choice(students).SID else "",
            str(job.DWZZJGDM) or "",
            "1"
        ]
        worksheet.append(row)

        # 保存文件
    workbook.save(file_path)

    print(f"数据已成功导出到文件: {file_path}")

    # df = pd.read_excel(file_path)
    # print(df.info())


def clear_old_click_jobs():
    """
    清除创建时间超过一年的 ClickJob 交互记录
    """
    # 获取当前时间
    now = timezone.now()
    # 计算一年前的时间
    one_year_ago = now - timedelta(days=365)

    # 查询创建时间超过一年的 ClickJob 记录
    old_click_jobs = ClickJob.objects.filter(create_time__lt=one_year_ago)

    # 删除这些记录
    count = old_click_jobs.count()  # 获取要删除的记录数量
    old_click_jobs.delete()  # 删除记录

    print(f"已删除 {count} 条创建时间超过一年的交互记录。")


"""
同步企业id和统一社会信用代码
"""


def syncCompanyCode():
    print("开始同步企业id和统一社会信用代码")
    key = '20240603380417974952305645404164571'
    secret = '567c49a47e435a09a27f33b6d18c91982886628a'
    access_token = ''
    per_page = 10000
    page = 1

    while True:
        # 获取token
        response = requests.get(f'https://dmp.xidian.edu.cn:8771/open_api/authentication/get_access_token?'
                                f'key={key}&secret={secret}')

        if response.status_code == 200:
            data = response.json()
            access_token = data.get('result').get('access_token')

        # 使用token发送请求
        response = requests.get(f'https://dmp.xidian.edu.cn:8771/open_api/customization/tgxxsdwxx/full?'
                                f'access_token={access_token}&per_page={per_page}&page={page}')
        if response.status_code == 200:
            data = response.json()
            max_page = data.get('result').get('max_page')
            print(page, max_page)

            codes = data.get('result').get('data')

            for code in codes:
                # if Student.objects.filter(SID=student.get('XH')).exists():
                #     continue

                saveStudent = CompanyCode()

                saveStudent.COMPANYID = code.get('WID')
                saveStudent.SHTYXYDM = code.get('SHTYXYDM')

                saveStudent.save()

            if page == max_page:
                break
            else:
                page = page + 1


def syncAllJobData():
    print("开始同步职位数据的全部字段")
    key = '20240603380417974952305645404164571'
    secret = '567c49a47e435a09a27f33b6d18c91982886628a'
    access_token = ''
    per_page = 10000
    page = 1

    while True:
        # 获取token
        response = requests.get(f'https://dmp.xidian.edu.cn:8771/open_api/authentication/get_access_token?'
                                f'key={key}&secret={secret}')

        if response.status_code == 200:
            data = response.json()
            access_token = data.get('result').get('access_token')

        # 使用token发送请求
        response = requests.get(f'https://dmp.xidian.edu.cn:8771/open_api/customization/tjyxtxxdviewjobinfo/full?'
                                f'access_token={access_token}&per_page={per_page}&page={page}')
        if response.status_code == 200:
            data = response.json()
            max_page = data.get('result').get('max_page')
            print(page, max_page)

            jobs = data.get('result').get('data')

            for job in jobs:
                if Job.objects.filter(DWZZJGDM=job.get('JOBID')).exists():
                    updateJob = Job.objects.get(DWZZJGDM=job.get('JOBID'))
                    # todo 更新字段信息
                    updateJob.NUM = job.get('NUM')
                    updateJob.save()

            if page == max_page:
                break
            else:
                page = page + 1


"""
设置定时任务
"""
# 1.实例化调度器
scheduler = BackgroundScheduler()
# 2.调度器使用DjangoJobStore()
scheduler.add_jobstore(DjangoJobStore(), "default")


# 3.设置定时任务
@register_job(scheduler, "interval", hours=24, replace_existing=True)
def my_job():
    logger.info("===定时任务开始执行===")
    # 同步学生数据
    syncStudentData()
    # 给所有学生创建账号
    createUsersForStudents()
    # 同步职位数据
    syncJobData()
    # 同步企业id和社会统一信用代码
    syncCompanyCode()
    logger.info("===定时任务执行完成===")


# 开启定时任务
# scheduler.start()

# 同步学生数据
# syncStudentData()
# studentToXlsx()

# 给所有学生创建账号
# createUsersForStudents()

# 同步职位数据
# syncJobData()
# jobToXlsx()
# 初始化交互记录
# createAction()


# 分析职位数据
# sync_job_data_and_analyze_update_time()

# 同步企业id和社会统一信用代码
# syncCompanyCode()


# 重新训练冷启动模型
# studentToXlsx()
# jobToXlsx()
# createAction()
# train()

# 同步职位数据的所有字段
# syncAllJobData()
